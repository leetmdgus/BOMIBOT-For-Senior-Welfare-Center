"""Excel·HWPX·DOCX → HTML 미리보기 (lxml / openpyxl / mammoth)."""

from __future__ import annotations

import csv
import html
import io
import re
import zipfile
from typing import Any

from lxml import etree

from app.application.hwpx.render.preview_theme import (
    HWPX_PREVIEW_THEME_CSS,
    OFFICE_HWPX_PREVIEW_CSS,
)

HP_NS = "http://www.hancom.co.kr/hwpml/2011/paragraph"
HP = f"{{{HP_NS}}}"
HP_P = f"{HP}p"
HP_T = f"{HP}t"
HP_TBL = f"{HP}tbl"
HP_TR = f"{HP}tr"
HP_TC = f"{HP}tc"

def _resolve_preview_body(data: bytes, filename: str) -> str:
    lower = (filename or "").lower()
    if lower.endswith(".hwpx"):
        return _render_hwpx_html(data, filename=filename)
    if lower.endswith(".csv"):
        return _render_csv_html(data)
    if lower.endswith((".xlsx", ".xls")):
        return _render_excel_html(data, lower.endswith(".xls"))
    if lower.endswith(".docx"):
        return _render_docx_html(data)
    if lower.endswith(".doc"):
        return (
            '<p class="office-preview-empty">'
            "구형 Word(.doc)은 미리보기를 지원하지 않습니다. "
            "DOCX로 변환 후 업로드해 주세요."
            "</p>"
        )
    return '<p class="office-preview-empty">미리보기를 지원하지 않는 형식입니다.</p>'


def render_document_preview_html(data: bytes, filename: str) -> str:
    lower = (filename or "").lower()
    is_hwpx = lower.endswith(".hwpx")
    body = _resolve_preview_body(data, filename)

    title = html.escape(filename or "문서")
    styles = _preview_styles(is_hwpx=is_hwpx)
    preview_class = "office-preview hwpx-like" if is_hwpx else "office-preview"
    return (
        f"<!DOCTYPE html><html lang=\"ko\"><head><meta charset=\"utf-8\"/>"
        f"<title>{title}</title>{styles}</head>"
        f'<body><div class="{preview_class}"><h2 class="office-preview-title">{title}</h2>{body}</div></body></html>'
    )


def render_document_preview_fragment(data: bytes, filename: str) -> str:
    """임베드용 HTML 조각 (스타일 + 본문)."""
    lower = (filename or "").lower()
    is_hwpx = lower.endswith(".hwpx")
    body = _resolve_preview_body(data, filename)
    styles = _preview_styles(is_hwpx=is_hwpx)
    preview_class = "office-preview hwpx-like" if is_hwpx else "office-preview"
    return f"{styles}<div class=\"{preview_class}\">{body}</div>"


def _preview_styles(*, is_hwpx: bool) -> str:
    if is_hwpx:
        return f"<style>{_OFFICE_BASE_STYLE}{HWPX_PREVIEW_THEME_CSS}{OFFICE_HWPX_PREVIEW_CSS}</style>"
    return f"<style>{_OFFICE_BASE_STYLE}</style>"


_OFFICE_BASE_STYLE = """
  .office-preview { font-family: "Malgun Gothic", "맑은 고딕", sans-serif; font-size: 13px; color: #111; line-height: 1.5; }
  .office-preview-title, .office-preview h2 { font-size: 15px; font-weight: 700; margin: 0 0 0.75rem; }
  .office-preview h3 { font-size: 14px; font-weight: 600; margin: 1rem 0 0.35rem; color: #334155; }
  .office-preview p { margin: 0.35rem 0; white-space: pre-wrap; }
  .office-preview-table { width: 100%; border-collapse: collapse; margin: 0.5rem 0 1rem; font-size: 12px; }
  .office-preview-table th, .office-preview-table td { border: 1px solid #cbd5e1; padding: 6px 8px; vertical-align: top; }
  .office-preview-table th { background: #f1f5f9; font-weight: 600; text-align: center; }
  .office-preview-empty { color: #64748b; padding: 2rem; text-align: center; }
  .office-preview-docx { max-width: 100%; }
  .office-preview-docx h1, .office-preview-docx h2, .office-preview-docx h3,
  .office-preview-docx h4, .office-preview-docx h5, .office-preview-docx h6 {
    margin: 0.75rem 0 0.35rem; font-weight: 700; line-height: 1.35;
  }
  .office-preview-docx ul, .office-preview-docx ol { margin: 0.35rem 0; padding-left: 1.5rem; }
  .office-preview-docx table { width: 100%; border-collapse: collapse; margin: 0.5rem 0 1rem; font-size: 12px; }
  .office-preview-docx table td, .office-preview-docx table th {
    border: 1px solid #cbd5e1; padding: 6px 8px; vertical-align: top;
  }
  .office-preview-docx img { max-width: 100%; height: auto; }
"""

_MAX_EXCEL_ROWS = 500
_MAX_EXCEL_COLS = 50


def is_previewable_filename(filename: str) -> bool:
    lower = (filename or "").lower()
    return lower.endswith((".xlsx", ".xls", ".csv", ".hwpx", ".docx", ".doc"))


def _render_csv_html(data: bytes) -> str:
    text = _decode_text(data)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return '<p class="office-preview-empty">내용이 없습니다.</p>'
    return _rows_to_table_html(rows[: _MAX_EXCEL_ROWS], header_row=True)


def _render_docx_html(data: bytes) -> str:
    try:
        import mammoth  # type: ignore[import-untyped]
    except ImportError:
        return (
            '<p class="office-preview-empty">'
            "Word(DOCX) 미리보기는 서버에 mammoth 패키지가 필요합니다."
            "</p>"
        )

    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            if "word/document.xml" not in archive.namelist():
                return '<p class="office-preview-empty">Word 문서 형식이 아닙니다.</p>'
    except zipfile.BadZipFile:
        return '<p class="office-preview-empty">DOCX 파일을 읽을 수 없습니다.</p>'

    result = mammoth.convert_to_html(io.BytesIO(data))
    body = (result.value or "").strip()
    if not body:
        return '<p class="office-preview-empty">내용이 없습니다.</p>'
    return f'<div class="office-preview-docx">{body}</div>'


def _render_excel_html(data: bytes, is_xls: bool) -> str:
    if is_xls:
        try:
            import xlrd  # type: ignore[import-untyped]
        except ImportError:
            return (
                '<p class="office-preview-empty">'
                "구형 Excel(.xls) 미리보기는 서버에 xlrd 패키지가 필요합니다."
                "</p>"
            )
        book = xlrd.open_workbook(file_contents=data)
        parts: list[str] = []
        for sheet in book.sheets():
            rows: list[list[str]] = []
            max_row = min(sheet.nrows, _MAX_EXCEL_ROWS)
            max_col = min(sheet.ncols, _MAX_EXCEL_COLS)
            for r in range(max_row):
                rows.append(
                    [
                        str(sheet.cell_value(r, c)) if c < sheet.ncols else ""
                        for c in range(max_col)
                    ]
                )
            parts.append(f"<h3>{html.escape(sheet.name)}</h3>")
            parts.append(_rows_to_table_html(rows, header_row=True))
        return "".join(parts) or '<p class="office-preview-empty">시트가 없습니다.</p>'

    try:
        from openpyxl import load_workbook
    except ImportError:
        return (
            '<p class="office-preview-empty">'
            "Excel 미리보기는 openpyxl 패키지가 필요합니다."
            "</p>"
        )

    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in book.worksheets:
        rows: list[list[str]] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= _MAX_EXCEL_ROWS:
                break
            cells = [
                _excel_cell_value(cell)
                for cell in list(row)[:_MAX_EXCEL_COLS]
            ]
            if any(cells):
                rows.append(cells)
        parts.append(f"<h3>{html.escape(sheet.title)}</h3>")
        parts.append(_rows_to_table_html(rows, header_row=True))
    book.close()
    return "".join(parts) or '<p class="office-preview-empty">시트가 없습니다.</p>'


def _excel_cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rows_to_table_html(rows: list[list[str]], *, header_row: bool) -> str:
    if not rows:
        return '<p class="office-preview-empty">표 데이터가 없습니다.</p>'

    parts = ['<table class="office-preview-table">']
    start = 0
    if header_row and rows:
        parts.append("<thead><tr>")
        for cell in rows[0]:
            parts.append(f"<th>{html.escape(cell)}</th>")
        parts.append("</tr></thead>")
        start = 1

    parts.append("<tbody>")
    for row in rows[start:]:
        parts.append("<tr>")
        for cell in row:
            parts.append(f"<td>{html.escape(cell)}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


def _guess_hwpx_template_kind(filename: str) -> str:
    name = filename or ""
    lower = name.lower()
    if "평가" in name or "evaluation" in lower:
        return "evaluation"
    return "plan"


def _render_hwpx_html(data: bytes, *, filename: str = "") -> str:
    """charPr·paraPr 반영 — render_json 파이프라인 (step3)."""
    try:
        from app.application.hwpx.render.file_json_render import (
            make_file_json_from_bytes,
            preview_from_file_json,
        )
        from app.application.hwpx.render.html_preview import render_json_to_body_fragment

        kind = _guess_hwpx_template_kind(filename)
        file_json = make_file_json_from_bytes(data, template_kind=kind)
        render_json = preview_from_file_json(file_json)
        paragraphs = (render_json.get("document") or {}).get("paragraphs") or []
        if paragraphs:
            return render_json_to_body_fragment(render_json)
    except Exception:
        import logging

        logging.getLogger(__name__).debug(
            "HWPX styled preview failed; falling back to plain extract",
            exc_info=True,
        )

    return _render_hwpx_html_plain(data)


def _render_hwpx_html_plain(data: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            section_names = sorted(
                name
                for name in archive.namelist()
                if name.replace("\\", "/").startswith("Contents/section")
                and name.lower().endswith(".xml")
            )
            if not section_names:
                return '<p class="office-preview-empty">본문 XML을 찾을 수 없습니다.</p>'

            parts: list[str] = []
            for index, name in enumerate(section_names):
                if len(section_names) > 1:
                    parts.append(f"<h3>구역 {index + 1}</h3>")
                root = etree.fromstring(archive.read(name))
                parts.extend(_hwpx_section_blocks(root))
            inner = "".join(parts) or '<p class="office-preview-empty">본문이 비어 있습니다.</p>'
            return f'<div class="hwpx-doc hwpx-doc--preview">{inner}</div>'
    except (zipfile.BadZipFile, etree.XMLSyntaxError):
        return '<p class="office-preview-empty">HWPX 파일을 읽을 수 없습니다.</p>'


def _is_inside_table(elem: etree._Element) -> bool:
    """표 셀(subList) 안 문단 — 별도 본문으로 렌더하지 않음."""
    parent = elem.getparent()
    while parent is not None:
        if parent.tag == HP_TBL:
            return True
        parent = parent.getparent()
    return False


def _hwpx_section_blocks(root: etree._Element) -> list[str]:
    blocks: list[str] = []
    for para in root.iter(HP_P):
        if _is_inside_table(para):
            continue

        table_parts = [_hwpx_table_html(tbl) for tbl in para.iter(HP_TBL)]
        if table_parts:
            blocks.extend(table_parts)
            continue

        text = _hp_text(para)
        if text.strip():
            blocks.append(f"<p>{html.escape(text)}</p>")

    for tbl in root.iter(HP_TBL):
        parent = tbl.getparent()
        if parent is not None and parent.tag == HP_P:
            continue
        blocks.append(_hwpx_table_html(tbl))

    return blocks


def _hp_text(elem: etree._Element) -> str:
    chunks: list[str] = []
    for node in elem.iter(HP_T):
        if node.text:
            chunks.append(node.text)
        if node.tail:
            chunks.append(node.tail)
    return "".join(chunks)


def _hwpx_cell_bg(tc: etree._Element, col_idx: int, col_cnt: int) -> str | None:
    ref = tc.get("borderFillIDRef")
    if ref == "4":
        return "#D9D9D9"
    if ref in {"2", "3", "5"}:
        return "#FFFFFF"
    if tc.get("header") == "1":
        return "#ECECEC"
    if col_cnt >= 2 and col_idx == 0:
        return "#D9D9D9"
    if col_cnt >= 2 and col_idx > 0:
        return "#FFFFFF"
    return None


def _hwpx_table_html(tbl: etree._Element) -> str:
    col_cnt = int(tbl.get("colCnt") or "0")
    rows_html: list[str] = []
    for tr in tbl.findall(HP_TR):
        cells_html: list[str] = []
        for col_idx, tc in enumerate(tr.findall(HP_TC)):
            text = _hp_text(tc) or " "
            is_header = tc.get("header") == "1"
            is_label_col = col_cnt >= 2 and col_idx == 0
            colspan = 1
            rowspan = 1
            span = tc.find(f"{HP}cellSpan")
            if span is not None:
                colspan = int(span.get("colSpan", "1"))
                rowspan = int(span.get("rowSpan", "1"))
            tag = "th" if is_header or is_label_col else "td"
            attrs = ""
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            if rowspan > 1:
                attrs += f' rowspan="{rowspan}"'
            css_class = ""
            if is_header or is_label_col:
                css_class = "hwpx-doc__label"
            elif col_cnt >= 2 and col_idx > 0:
                css_class = "hwpx-doc__value"
            bg = _hwpx_cell_bg(tc, col_idx, col_cnt)
            style_attr = f' style="background-color:{bg}"' if bg else ""
            class_attr = f' class="{css_class}"' if css_class else ""
            cells_html.append(
                f"<{tag}{attrs}{class_attr}{style_attr}>{html.escape(text)}</{tag}>"
            )
        if cells_html:
            rows_html.append(f"<tr>{''.join(cells_html)}</tr>")

    if not rows_html:
        return ""
    return (
        f'<div class="hwpx-doc__table-wrap">'
        f'<table class="hwpx-doc__table">{"".join(rows_html)}</table>'
        f"</div>"
    )


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")
